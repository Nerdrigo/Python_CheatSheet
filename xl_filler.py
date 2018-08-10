import openpyxl as xl

def fill_kyc(xl_file):
    '''
    This function has no purpose, it it only intended to remind me of how
    to fill an excell sheet 
    '''

    ####Basic things

    #Opening xl and choosing a seet
    wb = xl.load_workbook(xl_file)
    
    #Working on active sheet
    ws = wb.active
    #Chosing sheet by name
    ws = wb["Sheet Name"]

    #Filling xl sheet
    ws['A1'] = "Do or do not, there is no try"

    #Setting font
    ft = xl.styles.Font(size=8.5)
    a1 = ws['A1']
    a1.font = ft

    #Getting cell value
    #Empty cells will return None
    val = ws['A1'].value

    ######Adding rows to xl
    #Example list
    my_list = [2, 3, 5, 7]
    num_elem = len(my_list) 

    #available_space is whatever the number of available rows are
    #between "sections" in xl sheet
    available_space = 3

    #Offset is the number of spaces to add 
    #Negative numbers help identify how many spaces will be actually used
    num_offset = num_elem - available_space

    #If the number of elements exceeds the available space
    if num_offset > 0:
        #min_row = the row where the offset will take place
        #the offset occurs for rows below min_row
        ws._move_cells(min_row = 42, offset = num_offset, row_or_col="row")
    
    if num_elem > 0:
        for i,j in enumerate(range(42,42+available_space+admin_offset)):
            ws['A'+str(j)] = cliente.consejo[i]

    #Saving changes
    wb.save('filled_kyc.xlsx')